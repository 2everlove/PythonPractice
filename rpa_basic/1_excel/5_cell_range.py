from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

#Put data in line by line
ws.append(["No", "Eng", "Math"])

for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])


col_B = ws["B"] # get only Eng column
""" print(col_B)
for cell in col_B:
    print(cell.value)    """


col_range = ws["B:C"] # get Eng, Math column
""" for cols in col_range:
    for cell in cols:
        print(cell.value) """


row_title = ws[1] # get only 1st row
""" for cell in row_title:
    print(cell.value) """


row_range = ws[2:6] # get 2nd ~ 6th row
""" for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print() """


from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row] # get 2nd ~ last row
""" for rows in row_range:
    for cell in rows:
        #print(cell.value, end=" ")
        #print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        #print(xy, end=" ")
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print() """
        

#All rows
#print(ws.rows)  #!error <generator object Worksheet._cells_by_row at 0x000001FAE452C580>
#print(tuple(ws.rows)) #1st tuple: (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>) / 2nd : ...~
""" for row in tuple(ws.rows):
    print(row[1].value, end=" ") #Eng 16 54 50 6 3 56 69 79 21 56  """



#All columns
#print(tuple(ws.columns)) 
#1st tuple: 
# (<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>, <Cell 'Sheet'.A3>, <Cell 'Sheet'.A4>, 
# <Cell 'Sheet'.A5>, <Cell 'Sheet'.A6>, <Cell 'Sheet'.A7>, <Cell 'Sheet'.A8>, 
# <Cell 'Sheet'.A9>, <Cell 'Sheet'.A10>, <Cell 'Sheet'.A11>)
""" for column in tuple(ws.columns):
    print(column[0].value, end=" ") #No Eng Math """

""" for row in ws.iter_rows(): #All rows
    print(row[1].value) """


""" for column in ws.iter_cols(): #All column
    print(column[0].value) """

#2nd~11th row, 2nd~3rd column
""" for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
   # print(row[0].value, row[1].value) #Eng, Math
   print(row) #(<Cell 'Sheet'.B11>, <Cell 'Sheet'.C11>)"""


for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)#(<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>, <Cell 'Sheet'.A3>, <Cell 'Sheet'.A4>, <Cell 'Sheet'.A5>)

wb.save("sample.xlsx")