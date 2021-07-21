from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # Create new sheet which has default name
ws.title = "MySheet" # Change name of sheet
ws.sheet_properties.tabColor = "ff66ff" # When input RGB type, change color of tab

# [Sheet, MySheet, YourSheet]
ws1 = wb.create_sheet("YourSheet") # Create a sheet with a given name

# [Sheet, MySheet, NewSheet, YourSheet]
ws2 = wb.create_sheet("NewSheet", 2) # Create sheet in the 2nd index

new_ws = wb["NewSheet"] # Access the seat in a dictionary form

print(wb.sheetnames) # Check name of all sheets ['Sheet', 'MySheet', 'NewSheet', 'YourSheet']

# Copy sheet
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")