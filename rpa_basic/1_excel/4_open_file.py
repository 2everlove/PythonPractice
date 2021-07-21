from openpyxl import  load_workbook # Load File
wb = load_workbook("sample.xlsx") # Load wb from sample.xlsx
ws = wb.active # Activated sheet

# Load data from cell
""" for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 .. \n
    print() """

# When we don't know the number of cells
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 .. \n
    print()