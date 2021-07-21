from openpyxl import Workbook


wb = Workbook()
ws = wb.active
ws.title = "MikaSheet"

# Enter 1 in A1 cell 
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # Show a Information of A1 cell # <Cell 'MikaSheet'.A1>
print(ws["A1"].value) # Show a value of A1 cell # 1
print(ws["A10"].value) # Value is Null # None

# row = 1, 2, 3
# column = A(1), B(2), C(3)
print(ws.cell(column=1, row=1).value) # ws["A1"].value # 1
print(ws.cell(column=2, row=1).value) # ws["B1"].value # 4

c = ws.cell(column=3, row=1, value=10) # ws["C1"] = 10 # ws["C1"].value = 10
print(c.value) #ws["C1"].value # 10

from random import *

# Fill random numbers using repetitive statements # A1 ~ J10
index = 1
for x in range(1, 11): # Number of row is 10
    for y in range(1, 11): # Number of column is 10
        #ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100
        ws.cell(row=x, column=y, value=index)
        index += 1
        
wb.save("sample.xlsx")